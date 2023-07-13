# coding = utf-8
# Author: Hami
# Date: 2022/4/23 18:10

import openpyxl

class ReadFile:
    def __init__(self, fileName):
        self.fileName = fileName
        self.file = openpyxl.load_workbook(self.fileName)
        self.sheetnames = self.file.sheetnames
        self.sheet_data = self.file[self.sheetnames[0]]

    def read_excel(self):
        rows = self.sheet_data.max_row + 1
        cols = self.sheet_data.max_column + 1

        li_data = []

        for i in range(2, rows):
            info = {}
            for j in range(1, cols):
                value = self.sheet_data.cell(i, j).value
                # 字典名["key"] = value
                info[self.sheet_data.cell(1, j).value] = value
            li_data.append(info)

        return li_data

    def set_excel_values(self, i, j, values):
        self.sheet_data.cell(i, j).value = values
        self.file.save(self.fileName)


if __name__ == '__main__':
    rd = ReadFile(r"E:\python-learning\UT\data\Test_log.xlsx")
    print(rd.read_excel())
    # rd.set_excel_values(3, 1,"111")
